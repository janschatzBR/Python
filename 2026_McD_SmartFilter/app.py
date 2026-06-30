That is a great update for the interface! It makes the instructions much clearer for anyone using the tool.

Just a quick heads-up: in Python and Streamlit, the command for a new line is a backslash `\n` rather than a forward slash `/n`. However, since Streamlit supports Markdown, it's actually much cleaner to write this using a multi-line string (with triple quotes `"""`). This makes your code much easier to read and automatically formats the bullet points perfectly!

Here is the exact replacement for that section:

```python
# --- Web Interface (Streamlit) ---

st.set_page_config(page_title="US McOpCo Roster Changes", page_icon="📊", layout="wide")

st.title("US McOpCo Roster Changes")
st.markdown("""
Upload your **Previous** and **Current** 'McOpCo Alignment' spreadsheets. 

The system will compare the two to generate:
- Summary of changes
- Extraction of specific columns with **red text** from the current file
""")

```

To make sure your file stays perfectly intact, here is the complete, final `app.py` code with your new interface text seamlessly integrated at the bottom:

```python
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from copy import copy

def get_data_dictionary(ws, target_headers, check_format=False):
    """
    Helper function to extract sheet data into a dictionary keyed by National Store #.
    If check_format=True, it also evaluates the font color to flag red text.
    """
    normalized_targets = {" ".join(h.lower().split()): h for h in target_headers}
    found_columns_map = {}
    
    # Locate columns
    for col_idx in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=2, column=col_idx).value
        if not cell_val:
            cell_val = ws.cell(row=1, column=col_idx).value
            
        if cell_val:
            normalized_cell_text = " ".join(str(cell_val).lower().split())
            if normalized_cell_text in normalized_targets:
                found_columns_map[normalized_targets[normalized_cell_text]] = col_idx
                
    if "National Store #" not in found_columns_map:
        return (None, found_columns_map, None) if check_format else (None, found_columns_map)
        
    data_dict = {}
    red_dict = {} if check_format else None
    
    # Extract row values. If checking format, we must inspect cell objects instead of just values.
    for row in ws.iter_rows(min_row=3, values_only=not check_format):
        nsn_col_index = found_columns_map["National Store #"] - 1 # 0-indexed for tuple
        
        # Determine cell value depending on values_only flag
        nsn_val = row[nsn_col_index] if not check_format else row[nsn_col_index].value
        
        if nsn_val is None:
            continue
            
        nsn = str(nsn_val).strip()
        row_data = {}
        row_reds = set()
        
        for header in target_headers:
            if header in found_columns_map:
                val_idx = found_columns_map[header] - 1
                if val_idx < len(row):
                    if check_format:
                        cell = row[val_idx]
                        cell_val = cell.value
                        # Check if the font is red
                        if cell.font and cell.font.color and cell.font.color.type == 'rgb':
                            if cell.font.color.rgb in ['FF0000', 'FFFF0000']:
                                row_reds.add(header)
                    else:
                        cell_val = row[val_idx]
                        
                    row_data[header] = str(cell_val).strip() if cell_val is not None else "N/A"
                else:
                    row_data[header] = "N/A"
            else:
                row_data[header] = "N/A"
                
        data_dict[nsn] = row_data
        if check_format:
            red_dict[nsn] = row_reds
            
    if check_format:
        return data_dict, found_columns_map, red_dict
    return data_dict, found_columns_map

def process_excel(file_prev, file_curr):
    # Load workbooks
    wb_prev = openpyxl.load_workbook(file_prev)
    wb_curr = openpyxl.load_workbook(file_curr)
    
    if "Field Contacts" not in wb_prev.sheetnames or "Field Contacts" not in wb_curr.sheetnames:
        st.error("Error: Both files must contain a tab named 'Field Contacts'.")
        return None
        
    ws_prev = wb_prev["Field Contacts"]
    ws_curr = wb_curr["Field Contacts"]
    
    # All columns we need to extract from the spreadsheet
    target_headers = [
        "National Store #", "District", "PEOPLE PARTNERS", "LEX Supervisor", 
        "Risk", "Security", "VP", "Operations Officer", "Deployment Lead", 
        "Operations Manager", "Area Supervisor", "PC Ops Name"
    ]
    
    # ONLY the targeted roles that should trigger an "Action Required" log if changed
    roles_to_track = [
        "PEOPLE PARTNERS", "LEX Supervisor", "Risk", "Security", "VP", 
        "Operations Officer", "Deployment Lead", "Operations Manager", "Area Supervisor"
    ]
    
    # 1. Extract Data Dictionaries for Comparison
    prev_data, _ = get_data_dictionary(ws_prev, target_headers, check_format=False)
    
    curr_result = get_data_dictionary(ws_curr, target_headers, check_format=True)
    curr_data = curr_result[0]
    curr_cols_map = curr_result[1]
    curr_red_dict = curr_result[2]
    
    if prev_data is None or curr_data is None:
        st.error("Error: 'National Store #' column not found in one or both files.")
        return None

    # --- COMPARISON LOGIC ---
    prev_nsns = set(prev_data.keys())
    curr_nsns = set(curr_data.keys())
    
    added_nsns = curr_nsns - prev_nsns
    removed_nsns = prev_nsns - curr_nsns
    common_nsns = curr_nsns.intersection(prev_nsns)
    
    changes_list = []
    
    # Check Added
    for nsn in added_nsns:
        pc = curr_data[nsn].get("PC Ops Name", "N/A")
        changes_list.append(["ADD", nsn, pc, "National Store #", "N/A", nsn, "Yes"])
        
    # Check Removed
    for nsn in removed_nsns:
        pc = prev_data[nsn].get("PC Ops Name", "N/A")
        changes_list.append(["REMOVE", nsn, pc, "National Store #", nsn, "N/A", "Yes"])
        
    # Check Changed (Only checks the specific monitored roles)
    for nsn in common_nsns:
        for col in roles_to_track:
            old_val = prev_data[nsn].get(col, "N/A")
            new_val = curr_data[nsn].get(col, "N/A")
            is_red = col in curr_red_dict.get(nsn, set())
            
            # Trigger change log if the value changed OR if the cell is formatted red
            if old_val != new_val or is_red:
                pc = curr_data[nsn].get("PC Ops Name", "N/A")
                action = "Yes" if old_val != new_val else "Yes (Red Text)"
                changes_list.append(["CHANGE", nsn, pc, col, old_val, new_val, action])

    # --- CREATE NEW WORKBOOK ---
    new_wb = openpyxl.Workbook()
    
    # TAB 1: SUMMARY
    ws_summary = new_wb.active
    ws_summary.title = "Summary"
    bold_font = Font(bold=True)
    
    ws_summary.append(["US McOpCo Roster Changes – Update"])
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.append([])
    
    ws_summary.append(["Key Highlights"])
    ws_summary.cell(row=3, column=1).font = bold_font
    ws_summary.append([f"✅ {len(added_nsns)} New Stores Added"])
    ws_summary.append([f"⚠️ {len(removed_nsns)} Store Removed"])
    
    # Count only the 'CHANGE' rows for Role Changes highlight
    num_role_changes = len([c for c in changes_list if c[0] == 'CHANGE'])
    ws_summary.append([f"🔄 {num_role_changes} Changes Detected"])
    ws_summary.append([])
    
    # --- STORE CHANGES (Added/Removed) ---
    ws_summary.append(["Store Changes"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font
    
    ws_summary.append(["➕ Stores Added"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font
    for nsn in added_nsns:
        pc = curr_data[nsn].get('PC Ops Name', 'N/A')
        ws_summary.append([f"• National Store # {nsn} – PC Ops Name: {pc}"])
        
    ws_summary.append([])
    ws_summary.append(["➖ Stores Removed"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font
    for nsn in removed_nsns:
        pc = prev_data[nsn].get('PC Ops Name', 'N/A')
        ws_summary.append([f"• National Store # {nsn} – PC Ops Name: {pc}"])

    ws_summary.append([])
    
    # --- ACTION REQUIRED CHANGES ---
    ws_summary.append(["Action Required Changes"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = bold_font
    
    summary_headers = ["Change Type", "National Store #", "PC Ops Name", "Change", "Previous", "New", "Action"]
    ws_summary.append(summary_headers)
    
    header_row_idx = ws_summary.max_row
    for col_num in range(1, len(summary_headers) + 1):
        ws_summary.cell(row=header_row_idx, column=col_num).font = bold_font
        ws_summary.cell(row=header_row_idx, column=col_num).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        
    for change in changes_list:
        ws_summary.append(change)

    # TAB 2: CURRENT DATASET CHANGES (Original Red-Text Logic)
    ws_changes = new_wb.create_sheet("Current dataset changes")
    
    valid_columns = []
    headers = []
    
    for target in target_headers:
        if target in curr_cols_map:
            valid_columns.append(curr_cols_map[target])
            headers.append(target)
            
    ws_changes.append(headers)
    
    # Find rows starting from row 3 in the CURRENT file
    for row in ws_curr.iter_rows(min_row=3):
        is_red_row = False
        
        for col_idx in valid_columns:
            cell = ws_curr.cell(row=row[0].row, column=col_idx)
            if cell.font and cell.font.color and cell.font.color.type == 'rgb':
                if cell.font.color.rgb in ['FF0000', 'FFFF0000']:
                    is_red_row = True
                    break
                    
        # Extract the valid columns in the new exact order and preserve fonts
        if is_red_row:
            row_data = []
            cell_fonts = []
            
            for col_idx in valid_columns:
                source_cell = ws_curr.cell(row=row[0].row, column=col_idx)
                row_data.append(source_cell.value)
                cell_fonts.append(copy(source_cell.font) if source_cell.font else None)
                
            ws_changes.append(row_data)
            
            new_row_idx = ws_changes.max_row
            for i, font in enumerate(cell_fonts):
                if font:
                    ws_changes.cell(row=new_row_idx, column=i + 1).font = font
            
    # Save the new workbook to an in-memory byte stream
    output = BytesIO()
    new_wb.save(output)
    output.seek(0)
    
    return output

# --- Web Interface (Streamlit) ---

st.set_page_config(page_title="US McOpCo Roster Changes", page_icon="📊", layout="wide")

st.title("US McOpCo Roster Changes")
st.markdown("""
Upload your **Previous** and **Current** 'McOpCo Alignment' spreadsheets. 

The system will compare the two to generate:
- Summary of changes
- Extraction of specific columns with **red text** from the current file
""")

col1, col2 = st.columns(2)

with col1:
    file_previous = st.file_uploader("1️⃣ Upload PREVIOUS File (.xlsx)", type=["xlsx"])
    
with col2:
    file_current = st.file_uploader("2️⃣ Upload CURRENT File (.xlsx)", type=["xlsx"])

if file_previous is not None and file_current is not None:
    if st.button("Generate Comparison & Filter", type="primary"):
        with st.spinner("Analyzing changes and processing files..."):
            processed_file_stream = process_excel(file_previous, file_current)
            
        if processed_file_stream:
            st.success("Comparison complete! File processed successfully.")
            
            # Download button widget
            st.download_button(
                label="📥 Download Output File",
                data=processed_file_stream,
                file_name="Filtered_And_Compared_Contacts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
