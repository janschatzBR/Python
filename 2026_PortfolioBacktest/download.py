#!/usr/bin/env python3
"""Baixa preços/eventos, converte para USD e gera um backtest auditável.

Uso:
    python download.py
    python download.py --validate-only
    python download.py --daily

A fonte é Yahoo Finance via yfinance. Dividendos são brutos e datados pelo
ex-dividend date informado pela fonte. Impostos, retenções, corretagem, spread
de câmbio e slippage não são considerados.
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
import warnings
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
try:
    import yfinance as yf
except ImportError:  # permite executar --validate-only antes da instalação
    yf = None
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

import config


SOURCE_NAME = "Yahoo Finance via yfinance"
SOURCE_BASE_URL = "https://finance.yahoo.com/quote/{ticker}/history"
ZERO_EPSILON = 1e-12


@dataclass(frozen=True)
class AssetSpec:
    ticker: str
    asset_class: str
    name: str
    currency: str
    price_scale: float
    dividend_scale: float


class CollectionError(RuntimeError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Coleta dados históricos e gera historico_carteira.xlsx/.json."
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Valida configuração e tickers.csv sem acessar a internet.",
    )
    parser.add_argument(
        "--daily",
        action="store_true",
        help="Inclui preços e câmbio diários no Excel/JSON (arquivo maior).",
    )
    return parser.parse_args()


def setup_logging(output_dir: Path) -> logging.Logger:
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("carteira_backtest")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    file_handler = logging.FileHandler(output_dir / config.LOG_FILE, encoding="utf-8")
    file_handler.setFormatter(formatter)
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger


def normalize_date_index(index: Iterable[Any]) -> pd.DatetimeIndex:
    idx = pd.DatetimeIndex(pd.to_datetime(index))
    if idx.tz is not None:
        idx = idx.tz_localize(None)
    return idx.normalize()


def parse_date(value: str | None, field_name: str) -> pd.Timestamp:
    if value is None:
        return pd.Timestamp(date.today())
    try:
        return pd.Timestamp(value).normalize()
    except Exception as exc:
        raise ValueError(f"{field_name} inválida: {value!r}") from exc


def validate_config() -> tuple[pd.Timestamp, pd.Timestamp, Path, Path]:
    start = parse_date(config.START_DATE, "START_DATE")
    end = parse_date(config.END_DATE, "END_DATE")
    if end < start:
        raise ValueError("END_DATE não pode ser anterior a START_DATE.")
    if float(config.INITIAL_CAPITAL_USD) <= 0:
        raise ValueError("INITIAL_CAPITAL_USD deve ser positivo.")

    project_dir = Path(__file__).resolve().parent
    ticker_path = project_dir / config.TICKERS_FILE
    output_dir = project_dir / config.OUTPUT_FOLDER
    if not ticker_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {ticker_path}")
    return start, end, ticker_path, output_dir


def load_assets(ticker_path: Path) -> list[AssetSpec]:
    df = pd.read_csv(ticker_path, dtype=str).fillna("")
    required = {
        "Ticker",
        "Classe",
        "Nome",
        "Moeda",
        "EscalaPreco",
        "EscalaDividendo",
        "Incluir",
    }
    missing = required.difference(df.columns)
    if missing:
        raise ValueError(f"Colunas ausentes em tickers.csv: {sorted(missing)}")

    include_values = df["Incluir"].str.strip().str.lower()
    df = df[include_values.isin({"1", "sim", "yes", "true", "x"})].copy()
    if df.empty:
        raise ValueError("Nenhum ativo marcado para inclusão.")

    if df["Ticker"].str.strip().duplicated().any():
        duplicate = df.loc[df["Ticker"].str.strip().duplicated(), "Ticker"].tolist()
        raise ValueError(f"Tickers duplicados: {duplicate}")

    assets: list[AssetSpec] = []
    for row in df.itertuples(index=False):
        ticker = str(row.Ticker).strip().upper()
        currency = str(row.Moeda).strip().upper()
        if not ticker or not currency:
            raise ValueError("Ticker e Moeda são obrigatórios em todas as linhas incluídas.")
        try:
            price_scale = float(str(row.EscalaPreco).replace(",", "."))
            dividend_scale = float(str(row.EscalaDividendo).replace(",", "."))
        except ValueError as exc:
            raise ValueError(f"Escalas inválidas para {ticker}.") from exc
        if price_scale <= 0 or dividend_scale <= 0:
            raise ValueError(f"Escalas devem ser positivas para {ticker}.")
        assets.append(
            AssetSpec(
                ticker=ticker,
                asset_class=str(row.Classe).strip(),
                name=str(row.Nome).strip(),
                currency=currency,
                price_scale=price_scale,
                dividend_scale=dividend_scale,
            )
        )
    return assets


def history_with_fallback(
    ticker: str,
    start: pd.Timestamp,
    end_exclusive: pd.Timestamp,
    repair: bool,
) -> pd.DataFrame:
    if yf is None:
        raise CollectionError("yfinance não está instalado. Execute: pip install -r requirements.txt")
    obj = yf.Ticker(ticker)
    kwargs: dict[str, Any] = {
        "start": start.strftime("%Y-%m-%d"),
        "end": end_exclusive.strftime("%Y-%m-%d"),
        "interval": "1d",
        "auto_adjust": False,
        "actions": True,
        "repair": repair,
        "keepna": True,
        "timeout": int(config.REQUEST_TIMEOUT_SECONDS),
    }
    attempts = [
        (),
        ("timeout",),
        ("timeout", "repair"),
    ]
    last_exc: Exception | None = None
    for keys_to_remove in attempts:
        attempt = {k: v for k, v in kwargs.items() if k not in keys_to_remove}
        try:
            history = obj.history(**attempt)
            if history is None or history.empty:
                raise CollectionError(f"Nenhum histórico retornado para {ticker}.")
            history = history.copy()
            history.index = normalize_date_index(history.index)
            history = history[~history.index.duplicated(keep="last")].sort_index()
            return history
        except Exception as exc:  # yfinance lança diferentes classes por versão
            last_exc = exc
    raise CollectionError(f"Falha ao baixar {ticker}: {last_exc}")


def ensure_history_columns(history: pd.DataFrame) -> pd.DataFrame:
    result = history.copy()
    numeric_defaults = {
        "Open": np.nan,
        "High": np.nan,
        "Low": np.nan,
        "Close": np.nan,
        "Adj Close": np.nan,
        "Volume": 0.0,
        "Dividends": 0.0,
        "Stock Splits": 0.0,
        "Capital Gains": 0.0,
    }
    for column, default in numeric_defaults.items():
        if column not in result.columns:
            result[column] = default
        result[column] = pd.to_numeric(result[column], errors="coerce")
    result["Dividends"] = result["Dividends"].fillna(0.0)
    result["Stock Splits"] = result["Stock Splits"].fillna(0.0)
    result["Capital Gains"] = result["Capital Gains"].fillna(0.0)
    return result


def fx_pair_candidates(currency: str) -> list[tuple[str, bool]]:
    currency = currency.upper()
    return [(f"{currency}USD=X", False), (f"USD{currency}=X", True)]


def download_fx_series(
    currency: str,
    start: pd.Timestamp,
    end_exclusive: pd.Timestamp,
    cache: dict[str, pd.Series],
) -> pd.Series:
    currency = currency.upper()
    if currency == "USD":
        idx = pd.date_range(start, end_exclusive - pd.Timedelta(days=1), freq="D")
        return pd.Series(1.0, index=idx, name="FX_USD")
    if currency in cache:
        return cache[currency]

    errors: list[str] = []
    for pair, invert in fx_pair_candidates(currency):
        try:
            hist = history_with_fallback(pair, start - pd.Timedelta(days=10), end_exclusive, False)
            hist = ensure_history_columns(hist)
            close = hist["Close"].dropna().astype(float)
            if close.empty:
                raise CollectionError("série de fechamento vazia")
            if invert:
                close = 1.0 / close
            close.name = "FX_USD"
            cache[currency] = close
            return close
        except Exception as exc:
            errors.append(f"{pair}: {exc}")
    raise CollectionError(
        f"Não foi possível obter câmbio {currency}/USD. Tentativas: {'; '.join(errors)}"
    )


def align_fx(asset_dates: pd.DatetimeIndex, fx: pd.Series) -> pd.Series:
    fx = fx.copy()
    fx.index = normalize_date_index(fx.index)
    union = asset_dates.union(fx.index).sort_values()
    aligned = fx.reindex(union).ffill().bfill().reindex(asset_dates)
    if aligned.isna().any():
        raise CollectionError("Série de câmbio contém lacunas após alinhamento.")
    return aligned.astype(float)


def next_valid_close(close_usd: pd.Series, event_date: pd.Timestamp) -> tuple[pd.Timestamp, float]:
    remaining = close_usd.loc[close_usd.index >= event_date].dropna()
    if remaining.empty:
        raise CollectionError(f"Sem preço para reinvestir dividendo em {event_date.date()}.")
    return pd.Timestamp(remaining.index[0]), float(remaining.iloc[0])


def safe_float(value: Any) -> float | None:
    try:
        number = float(value)
        if math.isfinite(number):
            return number
    except (TypeError, ValueError):
        pass
    return None


def annualized_return(final_value: float, initial_value: float, years: float) -> float | None:
    if initial_value <= 0 or final_value < 0 or years <= 0:
        return None
    if final_value == 0:
        return -1.0
    return (final_value / initial_value) ** (1.0 / years) - 1.0


def process_asset(
    asset: AssetSpec,
    history: pd.DataFrame,
    fx_series: pd.Series,
    allocation_usd: float,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], pd.DataFrame, pd.DataFrame]:
    hist = ensure_history_columns(history)
    hist = hist.loc[hist.index >= pd.Timestamp(config.START_DATE)].copy()
    if hist.empty:
        raise CollectionError("Sem dados dentro do período solicitado.")

    price_columns = ["Open", "High", "Low", "Close", "Adj Close"]
    for col in price_columns:
        hist[f"{col} Native"] = hist[col] * asset.price_scale
    hist["Dividend Native"] = hist["Dividends"] * asset.dividend_scale

    fx_aligned = align_fx(hist.index, fx_series)
    hist["FX USD"] = fx_aligned
    for col in price_columns:
        hist[f"{col} USD"] = hist[f"{col} Native"] * hist["FX USD"]
    hist["Dividend USD"] = hist["Dividend Native"] * hist["FX USD"]

    valid_close = hist["Close USD"].dropna()
    if valid_close.empty:
        raise CollectionError("Sem preços de fechamento válidos.")

    purchase_date = pd.Timestamp(valid_close.index[0])
    purchase_price_usd = float(valid_close.iloc[0])
    if purchase_price_usd <= 0:
        raise CollectionError("Preço inicial inválido.")
    last_date = pd.Timestamp(valid_close.index[-1])
    last_price_usd = float(valid_close.iloc[-1])
    purchase_price_native = float(hist.loc[purchase_date, "Close Native"])
    last_price_native = float(hist.loc[last_date, "Close Native"])
    purchase_fx = float(hist.loc[purchase_date, "FX USD"])
    last_fx = float(hist.loc[last_date, "FX USD"])
    purchase_adj_usd = safe_float(hist.loc[purchase_date, "Adj Close USD"])
    last_adj_usd = safe_float(hist.loc[last_date, "Adj Close USD"])

    initial_shares = allocation_usd / purchase_price_usd
    shares_cash = initial_shares
    shares_reinvest = initial_shares
    dividends_cash_usd = 0.0
    dividends_reinvest_gross_usd = 0.0

    dividend_rows: list[dict[str, Any]] = []
    split_rows: list[dict[str, Any]] = []
    key_price_dates: set[pd.Timestamp] = {purchase_date, last_date}

    events = hist.loc[hist.index > purchase_date]
    for event_date, row in events.iterrows():
        event_date = pd.Timestamp(event_date)
        split_ratio = safe_float(row.get("Stock Splits")) or 0.0
        if abs(split_ratio) > ZERO_EPSILON and abs(split_ratio - 1.0) > ZERO_EPSILON:
            # O Close do Yahoo é ajustado por splits. Dividendos históricos também
            # são apresentados na mesma base de ações. Reaplicar o split às ações
            # produziria dupla contagem, especialmente em reverse splits de REITs.
            split_rows.append(
                {
                    "Data": event_date,
                    "Ticker": asset.ticker,
                    "Classe": asset.asset_class,
                    "Razao Split": split_ratio,
                    "Acoes Base (Caixa)": shares_cash,
                    "Acoes Base (Reinv.)": shares_reinvest,
                    "Tratamento": "Somente auditoria; não reaplicado (série ajustada por split)",
                    "Fonte": SOURCE_BASE_URL.format(ticker=asset.ticker),
                }
            )
            key_price_dates.add(event_date)

        dividend_usd = safe_float(row.get("Dividend USD")) or 0.0
        dividend_native = safe_float(row.get("Dividend Native")) or 0.0
        if dividend_usd > ZERO_EPSILON:
            key_price_dates.add(event_date)
            cash_shares_before = shares_cash
            reinvest_shares_before = shares_reinvest
            cash_received = cash_shares_before * dividend_usd
            reinvest_cash = reinvest_shares_before * dividend_usd
            dividends_cash_usd += cash_received
            dividends_reinvest_gross_usd += reinvest_cash

            reinvest_price_date, reinvest_price_usd = next_valid_close(hist["Close USD"], event_date)
            reinvested_shares = 0.0
            if config.REINVEST_DIVIDENDS and reinvest_price_usd > 0:
                reinvested_shares = reinvest_cash / reinvest_price_usd
                shares_reinvest += reinvested_shares
                key_price_dates.add(reinvest_price_date)

            dividend_rows.append(
                {
                    "Data Ex": event_date,
                    "Ticker": asset.ticker,
                    "Classe": asset.asset_class,
                    "Dividendo/Acao Nativo": dividend_native,
                    "Moeda": asset.currency,
                    "FX USD": float(row["FX USD"]),
                    "Dividendo/Acao USD": dividend_usd,
                    "Acoes Base (Caixa)": cash_shares_before,
                    "Dividendo Caixa USD": cash_received,
                    "Acoes Base (Reinv.)": reinvest_shares_before,
                    "Dividendo Bruto Reinv. USD": reinvest_cash,
                    "Data Preco Reinv.": reinvest_price_date,
                    "Preco Reinv. USD": reinvest_price_usd,
                    "Acoes Compradas": reinvested_shares,
                    "Acoes Apos Reinv.": shares_reinvest,
                    "Fonte": SOURCE_BASE_URL.format(ticker=asset.ticker),
                }
            )

    market_value_cash = shares_cash * last_price_usd
    total_cash = market_value_cash + dividends_cash_usd
    market_value_reinvest = shares_reinvest * last_price_usd
    value_via_adj_close: float | None = None
    return_via_adj_close: float | None = None
    difference_vs_adj_close: float | None = None
    difference_vs_adj_close_pct: float | None = None
    if (
        purchase_adj_usd is not None
        and last_adj_usd is not None
        and purchase_adj_usd > 0
        and last_adj_usd > 0
    ):
        value_via_adj_close = allocation_usd * last_adj_usd / purchase_adj_usd
        return_via_adj_close = value_via_adj_close / allocation_usd - 1.0
        difference_vs_adj_close = market_value_reinvest - value_via_adj_close
        difference_vs_adj_close_pct = difference_vs_adj_close / value_via_adj_close

    years = max((last_date - purchase_date).days / 365.2425, 1.0 / 365.2425)
    reconciliation_status = "OK"
    if difference_vs_adj_close_pct is not None and abs(difference_vs_adj_close_pct) > 0.03:
        reconciliation_status = "REVISAR (>3% vs Adj Close)"

    summary = {
        "Ticker": asset.ticker,
        "Classe": asset.asset_class,
        "Nome": asset.name,
        "Moeda Nativa": asset.currency,
        "Data Compra": purchase_date,
        "Preco Compra Nativo": purchase_price_native,
        "FX Compra USD": purchase_fx,
        "Preco Compra USD": purchase_price_usd,
        "Investimento Inicial USD": allocation_usd,
        "Acoes Iniciais": initial_shares,
        "Acoes Finais (Caixa)": shares_cash,
        "Data Final": last_date,
        "Preco Final Nativo": last_price_native,
        "FX Final USD": last_fx,
        "Preco Final USD": last_price_usd,
        "Valor Posicao USD": market_value_cash,
        "Dividendos Separados USD": dividends_cash_usd,
        "Total c/ Dividendos Caixa USD": total_cash,
        "Retorno Total Caixa %": total_cash / allocation_usd - 1.0,
        "CAGR Caixa %": annualized_return(total_cash, allocation_usd, years),
        "Yield on Cost Acumulado %": dividends_cash_usd / allocation_usd,
        "Acoes Finais (Reinv.)": shares_reinvest,
        "Dividendos Brutos Gerados Reinv. USD": dividends_reinvest_gross_usd,
        "Valor Final Reinvestido USD": market_value_reinvest,
        "Retorno Total Reinv. %": market_value_reinvest / allocation_usd - 1.0,
        "CAGR Reinv. %": annualized_return(market_value_reinvest, allocation_usd, years),
        "Valor via Adj Close USD": value_via_adj_close,
        "Retorno via Adj Close %": return_via_adj_close,
        "Diferenca Reinv. vs Adj Close USD": difference_vs_adj_close,
        "Diferenca Reinv. vs Adj Close %": difference_vs_adj_close_pct,
        "Anos": years,
        "Fonte": SOURCE_BASE_URL.format(ticker=asset.ticker),
        "Status": reconciliation_status,
    }

    key_dates = sorted(d for d in key_price_dates if d in hist.index)
    key = hist.loc[key_dates].copy()
    key.insert(0, "Data", key.index)
    key.insert(1, "Ticker", asset.ticker)
    key.insert(2, "Classe", asset.asset_class)
    key.insert(3, "Moeda", asset.currency)
    key = key[
        [
            "Data",
            "Ticker",
            "Classe",
            "Moeda",
            "Open Native",
            "High Native",
            "Low Native",
            "Close Native",
            "Adj Close Native",
            "FX USD",
            "Close USD",
            "Adj Close USD",
            "Dividend Native",
            "Dividend USD",
            "Stock Splits",
            "Volume",
        ]
    ]

    daily = hist.copy()
    daily.insert(0, "Data", daily.index)
    daily.insert(1, "Ticker", asset.ticker)
    daily.insert(2, "Classe", asset.asset_class)
    daily.insert(3, "Moeda", asset.currency)
    daily = daily[
        [
            "Data",
            "Ticker",
            "Classe",
            "Moeda",
            "Open Native",
            "High Native",
            "Low Native",
            "Close Native",
            "Adj Close Native",
            "FX USD",
            "Open USD",
            "High USD",
            "Low USD",
            "Close USD",
            "Adj Close USD",
            "Volume",
            "Dividend Native",
            "Dividend USD",
            "Stock Splits",
            "Capital Gains",
        ]
    ]
    return summary, dividend_rows, split_rows, key, daily


def records_for_json(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df.empty:
        return []
    copy = df.copy()
    for col in copy.columns:
        if pd.api.types.is_datetime64_any_dtype(copy[col]):
            copy[col] = copy[col].dt.strftime("%Y-%m-%d")
    copy = copy.replace({np.nan: None, np.inf: None, -np.inf: None})
    return copy.to_dict(orient="records")


def build_assumptions(
    start: pd.Timestamp,
    requested_end: pd.Timestamp,
    asset_count: int,
    allocation: float,
) -> pd.DataFrame:
    rows = [
        ("Capital inicial", float(config.INITIAL_CAPITAL_USD), "USD"),
        ("Numero de ativos", asset_count, "SPCX excluido; VICI incluido"),
        ("Aporte por ativo", allocation, "USD; pesos iguais; fracoes permitidas"),
        ("Data alvo inicial", start, "Primeiro fechamento disponivel em ou apos a data"),
        ("Data final solicitada", requested_end, "Ultimo fechamento disponivel ate a data"),
        ("Preco", "Close ajustado por splits; não por dividendos", "Adj Close é mantido apenas para auditoria"),
        ("Dividendos", "Brutos e na base ajustada por splits", "Sem impostos, retencoes, taxas ou spread"),
        ("Splits", "Registrados, mas não reaplicados", "Evita dupla contagem porque a série Close já é ajustada por splits"),
        ("Checagem", "Adj Close", "Diferenças acima de 3% são marcadas para revisão; pode indicar eventos corporativos ou metodologia distinta"),
        ("Data do dividendo", "Ex-dividend date", "Conforme Yahoo Finance"),
        (
            "Reinvestimento",
            "Fechamento do dia ex-dividendo",
            "Se indisponivel, primeiro fechamento posterior",
        ),
        ("Moeda-base", "USD", "VWCE convertido de EUR; VHYL convertido de GBP"),
                ("Fonte", SOURCE_NAME, "Dados devem ser conferidos antes de decisoes financeiras"),
        ("Versao pandas", pd.__version__, "Ambiente de execução"),
        ("Versao yfinance", getattr(yf, "__version__", "não instalado"), "Ambiente de execução"),
        ("Gerado em", pd.Timestamp.now().floor("s"), "Horario local do computador"),
    ]
    return pd.DataFrame(rows, columns=["Premissa", "Valor", "Observacao"])


def add_total_row(summary: pd.DataFrame) -> pd.DataFrame:
    if summary.empty:
        return summary
    total: dict[str, Any] = {col: None for col in summary.columns}
    total["Ticker"] = "TOTAL"
    total["Classe"] = "Carteira"
    total["Nome"] = f"{len(summary)} ativos"
    sum_columns = [
        "Investimento Inicial USD",
        "Valor Posicao USD",
        "Dividendos Separados USD",
        "Total c/ Dividendos Caixa USD",
        "Dividendos Brutos Gerados Reinv. USD",
        "Valor Final Reinvestido USD",
        "Valor via Adj Close USD",
    ]
    for col in sum_columns:
        if col == "Valor via Adj Close USD" and not summary[col].notna().all():
            total[col] = None
        else:
            total[col] = float(summary[col].sum())
    initial = total["Investimento Inicial USD"]
    total_cash = total["Total c/ Dividendos Caixa USD"]
    reinvested = total["Valor Final Reinvestido USD"]
    total["Retorno Total Caixa %"] = total_cash / initial - 1.0
    total["Yield on Cost Acumulado %"] = total["Dividendos Separados USD"] / initial
    total["Retorno Total Reinv. %"] = reinvested / initial - 1.0
    adj_total = total.get("Valor via Adj Close USD")
    if adj_total is not None and adj_total > 0:
        total["Retorno via Adj Close %"] = adj_total / initial - 1.0
        total["Diferenca Reinv. vs Adj Close USD"] = reinvested - adj_total
        total["Diferenca Reinv. vs Adj Close %"] = (reinvested - adj_total) / adj_total
    # CAGR agregado usa o intervalo comum solicitado; diferenças de último pregão são mínimas.
    years = max(float(summary["Anos"].median()), 1.0 / 365.2425)
    total["Anos"] = years
    total["CAGR Caixa %"] = annualized_return(total_cash, initial, years)
    total["CAGR Reinv. %"] = annualized_return(reinvested, initial, years)
    total["Status"] = "OK"
    result = summary.copy()
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.*",
            category=FutureWarning,
        )
        result.loc[len(result)] = total
    return result


def write_excel(
    path: Path,
    summary: pd.DataFrame,
    dividends: pd.DataFrame,
    splits: pd.DataFrame,
    key_prices: pd.DataFrame,
    assumptions: pd.DataFrame,
    errors: pd.DataFrame,
    daily_prices: pd.DataFrame | None,
    daily_fx: pd.DataFrame | None,
) -> None:
    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        summary.to_excel(writer, sheet_name="Resumo", index=False)
        dividends.to_excel(writer, sheet_name="Dividendos", index=False)
        splits.to_excel(writer, sheet_name="Splits", index=False)
        key_prices.to_excel(writer, sheet_name="Cotacoes_Chave", index=False)
        assumptions.to_excel(writer, sheet_name="Premissas", index=False)
        errors.to_excel(writer, sheet_name="Erros", index=False)
        if daily_prices is not None:
            daily_prices.to_excel(writer, sheet_name="Precos_Diarios", index=False)
        if daily_fx is not None:
            daily_fx.to_excel(writer, sheet_name="FX_Diario", index=False)
    style_excel(path)


def style_excel(path: Path) -> None:
    wb = load_workbook(path)
    dark_fill = PatternFill("solid", fgColor="17365D")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    white_bold = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")
    top_border = Border(top=Side(style="thin", color="17365D"))

    percent_keywords = ("%", "CAGR", "Yield")
    share_keywords = ("Acoes", "Razao")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = dark_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 34

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            header = str(ws.cell(1, col_idx).value or "")
            max_len = min(
                max(len(str(c.value)) if c.value is not None else 0 for c in column_cells),
                42,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 42))
            for cell in list(column_cells)[1:]:
                cell.border = Border(bottom=thin_gray)
                if any(key in header for key in percent_keywords):
                    cell.number_format = '0.00%;[Red](0.00%);-'
                elif "USD" in header:
                    cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
                elif "FX" in header:
                    cell.number_format = '0.000000;[Red](0.000000);-'
                elif any(key in header for key in share_keywords):
                    cell.number_format = '0.000000;[Red](0.000000);-'
                elif "Data" in header or header == "Gerado em":
                    cell.number_format = "yyyy-mm-dd"
                elif "Preco" in header or "Dividendo" in header:
                    cell.number_format = '#,##0.0000;[Red](#,##0.0000);-'
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.0000;[Red](#,##0.0000);-'

        if ws.title == "Resumo":
            last_row = ws.max_row
            if last_row >= 2 and ws.cell(last_row, 1).value == "TOTAL":
                for cell in ws[last_row]:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                    cell.border = top_border
            headers = {cell.value: cell.column for cell in ws[1]}
            for header in ("Retorno Total Caixa %", "Retorno Total Reinv. %"):
                col = headers.get(header)
                if col:
                    rng = f"{get_column_letter(col)}2:{get_column_letter(col)}{max(2, last_row - 1)}"
                    ws.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="min",
                            start_color="F8696B",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FFEB84",
                            end_type="max",
                            end_color="63BE7B",
                        ),
                    )
        if ws.title == "Premissas":
            for row in range(2, ws.max_row + 1):
                ws.cell(row, 2).fill = warning_fill

    wb.save(path)


def main() -> int:
    args = parse_args()
    try:
        start, requested_end, ticker_path, output_dir = validate_config()
        logger = setup_logging(output_dir)
        assets = load_assets(ticker_path)
        allocation = float(config.INITIAL_CAPITAL_USD) / len(assets)

        logger.info(f"Ativos incluídos: {len(assets)}")
        logger.info(f"Capital por ativo: US$ {allocation:,.6f}")
        logger.info(f"Período alvo: {start.date()} a {requested_end.date()}")
        if args.validate_only:
            logger.info("Validação concluída. Nenhum download foi executado.")
            return 0

        end_exclusive = requested_end + pd.Timedelta(days=1)
        save_daily = bool(args.daily or config.SAVE_DAILY_PRICES)
        save_fx = bool(args.daily or config.SAVE_DAILY_FX)

        fx_cache: dict[str, pd.Series] = {}
        summaries: list[dict[str, Any]] = []
        dividends_all: list[dict[str, Any]] = []
        splits_all: list[dict[str, Any]] = []
        key_frames: list[pd.DataFrame] = []
        daily_frames: list[pd.DataFrame] = []
        errors: list[dict[str, Any]] = []

        for asset in tqdm(assets, desc="Baixando ativos", unit="ativo"):
            try:
                logger.info(f"Processando {asset.ticker}...")
                history = history_with_fallback(
                    asset.ticker,
                    start,
                    end_exclusive,
                    bool(config.YFINANCE_REPAIR),
                )
                fx = download_fx_series(asset.currency, start, end_exclusive, fx_cache)
                summary, div_rows, split_rows, key, daily = process_asset(
                    asset, history, fx, allocation
                )
                summaries.append(summary)
                dividends_all.extend(div_rows)
                splits_all.extend(split_rows)
                key_frames.append(key)
                if save_daily:
                    daily_frames.append(daily)
                logger.info(f"{asset.ticker}: OK")
            except Exception as exc:
                message = str(exc)
                logger.error(f"{asset.ticker}: ERRO - {message}")
                errors.append(
                    {
                        "Ticker": asset.ticker,
                        "Classe": asset.asset_class,
                        "Nome": asset.name,
                        "Erro": message,
                        "Fonte": SOURCE_BASE_URL.format(ticker=asset.ticker),
                    }
                )

        summary_df = pd.DataFrame(summaries)
        if not summary_df.empty:
            class_order = {asset.ticker: idx for idx, asset in enumerate(assets)}
            summary_df["_ordem"] = summary_df["Ticker"].map(class_order)
            summary_df = summary_df.sort_values("_ordem").drop(columns="_ordem")
            summary_df = add_total_row(summary_df)

        dividends_df = pd.DataFrame(dividends_all)
        if not dividends_df.empty:
            dividends_df = dividends_df.sort_values(["Data Ex", "Ticker"]).reset_index(drop=True)
        splits_df = pd.DataFrame(splits_all)
        if not splits_df.empty:
            splits_df = splits_df.sort_values(["Data", "Ticker"]).reset_index(drop=True)
        key_df = pd.concat(key_frames, ignore_index=True) if key_frames else pd.DataFrame()
        if not key_df.empty:
            key_df = key_df.sort_values(["Ticker", "Data"]).reset_index(drop=True)
        daily_df = pd.concat(daily_frames, ignore_index=True) if daily_frames else None
        if daily_df is not None and not daily_df.empty:
            daily_df = daily_df.sort_values(["Ticker", "Data"]).reset_index(drop=True)

        errors_df = pd.DataFrame(errors, columns=["Ticker", "Classe", "Nome", "Erro", "Fonte"])
        assumptions_df = build_assumptions(start, requested_end, len(assets), allocation)

        fx_df: pd.DataFrame | None = None
        if save_fx:
            fx_rows: list[pd.DataFrame] = []
            for currency, series in fx_cache.items():
                frame = series.rename("FX USD").to_frame().reset_index(names="Data")
                frame.insert(1, "Moeda", currency)
                fx_rows.append(frame)
            fx_df = pd.concat(fx_rows, ignore_index=True) if fx_rows else pd.DataFrame()

        excel_path = output_dir / config.OUTPUT_EXCEL
        json_path = output_dir / config.OUTPUT_JSON
        write_excel(
            excel_path,
            summary_df,
            dividends_df,
            splits_df,
            key_df,
            assumptions_df,
            errors_df,
            daily_df,
            fx_df,
        )

        payload = {
            "metadata": {
                "generated_at": pd.Timestamp.now().isoformat(),
                "start_date": start.strftime("%Y-%m-%d"),
                "requested_end_date": requested_end.strftime("%Y-%m-%d"),
                "initial_capital_usd": float(config.INITIAL_CAPITAL_USD),
                "asset_count_requested": len(assets),
                "asset_count_ok": len(summaries),
                "asset_count_errors": len(errors),
                "allocation_per_asset_usd": allocation,
                "source": SOURCE_NAME,
                "pandas_version": pd.__version__,
                "yfinance_version": getattr(yf, "__version__", None),
                "daily_prices_included": save_daily,
                "daily_fx_included": save_fx,
            },
            "summary": records_for_json(summary_df),
            "dividends": records_for_json(dividends_df),
            "splits": records_for_json(splits_df),
            "key_prices": records_for_json(key_df),
            "errors": records_for_json(errors_df),
        }
        if save_daily and daily_df is not None:
            payload["daily_prices"] = records_for_json(daily_df)
        if save_fx and fx_df is not None:
            payload["daily_fx"] = records_for_json(fx_df)

        with json_path.open("w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)

        logger.info("")
        logger.info("Concluído.")
        logger.info(f"Ativos OK: {len(summaries)}")
        logger.info(f"Ativos com erro: {len(errors)}")
        logger.info(f"Excel: {excel_path}")
        logger.info(f"JSON: {json_path}")
        logger.info(f"Log: {output_dir / config.LOG_FILE}")
        if errors:
            logger.info("Consulte a aba 'Erros' e o log antes de usar os resultados.")
        return 0 if summaries else 2
    except Exception as exc:
        print(f"ERRO FATAL: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
